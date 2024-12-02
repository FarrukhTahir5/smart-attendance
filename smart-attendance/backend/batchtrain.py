import os
import cv2
import numpy as np
import pickle

from mtcnn import MTCNN
from keras_facenet import FaceNet
from openpyxl import load_workbook
from datetime import datetime

mtcnn = MTCNN()
facenet = FaceNet()

def extract_faces(image_path):
    try:
        img = cv2.imread(image_path)
        if img is None:
            raise Exception("Unable to load image.")
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        faces = mtcnn.detect_faces(img_rgb)
        if faces:
            extracted_faces = []
            for face in faces:
                x, y, w, h = face['box']
                face_img = img_rgb[y:y+h, x:x+w]
                extracted_faces.append((face_img, (x, y, x+w, y+h)))  # Store face image and bounding box coordinates
            return extracted_faces
        else:
            return None
    except Exception as e:
        print(f"Error: {e}")
        return None

def recognize_faces(unknown_face_encoding, stored_embeddings, rollno, threshold=0.92):
    min_distance = float("inf")
    identity = None
    for i, known_face_encoding in enumerate(stored_embeddings):
        distance = np.linalg.norm(unknown_face_encoding - known_face_encoding)
        if distance < min_distance:
            min_distance = distance
            identity = rollno[i]
    if min_distance <= threshold:
        return identity
    else:
        return "Unknown"
def load_known_faces(known_faces_dir):
    stored_embeddings = []
    rollno = []
    
    # Check if saved encodings file exists
    encodings_file = "stored_embeddings.pkl"
    
    if os.path.exists(encodings_file):
        with open(encodings_file, 'rb') as f:
            stored_embeddings, rollno = pickle.load(f)
    else:
        # Iterate over each folder in the directory
        for folder_name in os.listdir(known_faces_dir):
            folder_path = os.path.join(known_faces_dir, folder_name)
            if os.path.isdir(folder_path):
                # Collect image paths for the person
                image_paths = [os.path.join(folder_path, image_name) for image_name in os.listdir(folder_path)]
                # Load multiple images for the person
                for image_path in image_paths:
                    # Extract faces from the image
                    faces = extract_faces(image_path)
                    if faces:
                        # Encode faces using FaceNet
                        face_encodings = facenet.embeddings([face[0] for face in faces])
                        # Store face encodings and names
                        for i, face_encoding in enumerate(face_encodings):
                            stored_embeddings.append(face_encoding)
                            rollno.append(folder_name)
        
        # Save encodings to file
        with open(encodings_file, 'wb') as f:
            pickle.dump((stored_embeddings, rollno), f)
            
    return stored_embeddings, rollno


def draw_boxes(image, faces, recognized_friends):
    print(faces[0])
    for i, (face_img, bbox) in enumerate(faces):
        x1, y1, x2, y2 = bbox
        # Draw bounding box
        cv2.rectangle(image, (x1, y1), (x2, y2), (0, 255, 0), 2)
        # Label recognized friend
        cv2.putText(image, recognized_friends[i], (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (36,255,12), 2)
    return image

train_folder = "./storage/batches/32/ai"

# Load known faces
stored_embeddings, rollno = load_known_faces(train_folder)
print(rollno)

group_image_path = "./group.jpg"
group_faces = extract_faces(group_image_path)
print(group_faces[0])
recognized_friends = []
if group_faces:
    # Recognize faces
    for face_img, _ in group_faces:
        face_encodings = facenet.embeddings([face_img])
        if len(face_encodings) > 0:
            recognized_friend = recognize_faces(face_encodings[0], stored_embeddings, rollno)
            recognized_friends.append(recognized_friend)
        else:
            recognized_friends.append("Unknown")

    # Draw bounding boxes and labels
    group_image = cv2.imread(group_image_path)
    annotated_image = draw_boxes(group_image.copy(), group_faces, recognized_friends)
    annotated_image_path = "./annotated_image.jpg"
    cv2.imwrite(annotated_image_path, annotated_image)
    print(f"Annotated image saved as {annotated_image_path}")
    attendance_file = "./attendance.xlsx"
    wb = load_workbook(attendance_file)
    ws = wb.active

    # Get today's date
    today_date = datetime.today().strftime('%Y-%m-%d')

    # Check if today's date column exists, if not, add it
    if today_date not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=ws.max_column+1, value=today_date)

    # Loop through recognized friends and mark their attendance
    # Loop through recognized friends and mark their attendance
    for name in recognized_friends:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == name:
                # Find the cell corresponding to today's date and mark attendance
                today_column_index = None
                for i, cell in enumerate(ws[1]):
                    if cell.value == today_date:
                        today_column_index = i + 1
                        break
                if today_column_index is not None:
                    attendance_cell = ws.cell(row=row[0].row, column=today_column_index)
                    attendance_cell.value = "Present"
                    break


    # Save changes to Excel file
    wb.save(attendance_file)
    print(f"Attendance saved to {attendance_file}")
else:
    print("No faces detected in the group image.")