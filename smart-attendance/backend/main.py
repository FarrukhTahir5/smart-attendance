from fastapi import FastAPI, UploadFile, File, Form
from typing import List, Optional
import os
import cv2
import numpy as np
from mtcnn import MTCNN
from keras_facenet import FaceNet
from datetime import datetime
import pickle
import json
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook,load_workbook
import datetime

from fastapi.staticfiles import StaticFiles


app = FastAPI()

# Initialize MTCNN and FaceNet
mtcnn = MTCNN()
facenet = FaceNet()

# Base storage directory
BASE_DIR = "storage"
os.makedirs(BASE_DIR, exist_ok=True)

# Serve the "attendance" directory as static files
attendance_static_dir = os.path.join(BASE_DIR, "attendance")
os.makedirs(attendance_static_dir, exist_ok=True)
app.mount("/static", StaticFiles(directory=attendance_static_dir), name="static")

@app.get("/dropdown-data")
async def dropdown_data():
    data = {
        "departments": {
            "Computer Science": {
                "programs": {
                    "BS CS": ["Fall 2020", "Spring 2021"],
                    "MS CS": ["Spring 2021"]
                },
                "classes": {
                    "Fall 2020": [
                        "CS101 - Group A - Dr. Smith",
                        "CS101 - Group B - Dr. Jane"
                    ],
                    "Spring 2021": [
                        "CS102 - Group A - Dr. John"
                    ]
                }
            },
            "Electrical Engineering": {
                "programs": {
                    "BS EE": ["Fall 2020"],
                    "MS EE": ["Spring 2021"]
                },
                "classes": {
                    "Fall 2020": ["EE101 - Group A - Dr. Brown"],
                    "Spring 2021": ["EE102 - Group A - Dr. Alice"]
                }
            }
        }
    }
    return data

# Add CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins; use specific origins in production
    allow_credentials=True,
    allow_methods=["*"],  # Allow all HTTP methods
    allow_headers=["*"],  # Allow all headers
)

@app.post("/upload-student-image")
async def upload_student_image(
    name: str = Form(...),
    rollno: str = Form(...),
    program: str = Form(...),
    batch_number: str = Form(...),
    file: UploadFile = File(...)
):        
    # Create directory for the batch and program
    student_dir = f"{BASE_DIR}/{"batches"}/{batch_number}/{program}/{rollno}"
    photos_dir = f"{student_dir}/photos"
    os.makedirs(photos_dir, exist_ok=True)

    # Process each uploaded file
    file_path = f"{photos_dir}/{file.filename}"
    with open(file_path, "wb") as f:
        f.write(await file.read())
   
    return {
        "message": "file uploaded"
    }


# Helper function to update Excel attendance sheet
def update_attendance_excel(class_dir, recognized_students):
    attendance_file = f"{class_dir}/attendance.xlsx"
    today_date = datetime.today().strftime('%Y-%m-%d')

    if not os.path.exists(attendance_file):
        # Create a new Excel file if it doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.cell(row=1, column=1, value="Roll No")
        ws.cell(row=1, column=2, value=today_date)
        wb.save(attendance_file)

    # Load the Excel file
    wb = load_workbook(attendance_file)
    ws = wb.active

    # Add today's date column if not already present
    if today_date not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=ws.max_column + 1, value=today_date)

    # Mark attendance for recognized students
    today_column_index = None
    for i, cell in enumerate(ws[1]):
        if cell.value == today_date:
            today_column_index = i + 1
            break

    if today_column_index is not None:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value in recognized_students:
                ws.cell(row=row[0].row, column=today_column_index, value="Present")

    wb.save(attendance_file)
    return attendance_file


@app.post("/register")
async def register_student(
    name: str = Form(...),
    rollno: str = Form(...),
    batch_number: str = Form(...),
    program: str = Form(...),
    files: List[UploadFile] = File(...),
):
    try:
        # Log all received data
        print("Received data:", {
            "name": name, 
            "rollno": rollno, 
            "batch_number": batch_number, 
            "program": program,
            "files": [file.filename for file in files]
        })
        
        print(f"Received name: {name}, rollno: {rollno}, batch_number: {batch_number}, program: {program}")
        print(f"Received files: {[file.filename for file in files]}")
    except Exception as e:
        print("Error parsing request:", e)
        return {"error": "Invalid data received"}
    
    # Create directory for the batch and program
    student_dir = f"{BASE_DIR}/batches/{batch_number}/{program}/{rollno}"
    photos_dir = f"{student_dir}/photos"
    os.makedirs(photos_dir, exist_ok=True)

    # This will store the embeddings and roll number pairs
    stored_embeddings = []
    rollnos = []

    # Process each uploaded file
    for file in files:
        file_path = f"{photos_dir}/{file.filename}"
        with open(file_path, "wb") as f:
            f.write(await file.read())

        # Detect faces and compute embeddings
        image = cv2.imread(file_path)
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        faces = mtcnn.detect_faces(image_rgb)

        if not faces:
            return {"error": f"No face detected in {file.filename}"}

        for face in faces:
            x, y, w, h = face["box"]
            face_crop = image_rgb[y:y+h, x:x+w]
            embedding = facenet.embeddings([face_crop])[0]
            stored_embeddings.append(embedding)
            rollnos.append(rollno)  # Store the roll number for each embedding

    # Save individual embeddings for the student in the same format
    embeddings_path = f"{student_dir}/embeddings.pkl"
    with open(embeddings_path, "wb") as f:
        pickle.dump((stored_embeddings, rollnos), f)

    # Update batch-level embeddings
    batch_embeddings_path = f"{BASE_DIR}/batches/{batch_number}/{program}/batch_embeddings.pkl"
    all_embeddings = {}

    if os.path.exists(batch_embeddings_path):
        with open(batch_embeddings_path, "rb") as f:
            all_embeddings = pickle.load(f)

    # Add the student's embeddings to the batch-level embeddings
    all_embeddings[rollno] = stored_embeddings

    # Save the updated batch embeddings
    with open(batch_embeddings_path, "wb") as f:
        pickle.dump(all_embeddings, f)

    return {"message": f"Student {name} (Roll No: {rollno}) registered successfully in Batch {batch_number} for {program}."}




def extract_faces(image):
    """Extract faces from the uploaded image using MTCNN."""
    try:
        img_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        faces = mtcnn.detect_faces(img_rgb)
        extracted_faces = []
        for face in faces:
            x, y, w, h = face['box']
            face_img = img_rgb[y:y+h, x:x+w]
            extracted_faces.append((face_img, (x, y, x+w, y+h)))  # Store face image and bounding box coordinates
        return extracted_faces
    except Exception as e:
        print(f"Error during face extraction: {e}")
        return []

def recognize_faces(embedding, class_embeddings, threshold=0.92):
    """Compare the face embedding with class embeddings and return recognized student ID."""
    print(len(class_embeddings))

    # Assuming class_embeddings is a tuple of two lists
    embeddings_list, names_list = class_embeddings

    # Now iterate over the embeddings and names
    for rollno, stored_embedding in zip(names_list, embeddings_list):
       distance = np.linalg.norm(embedding - stored_embedding)
       if distance < threshold:  # Recognition threshold
           return rollno
    return "Unknown"


def update_attendance_excel(class_dir, recognized_students):
    """Update the attendance Excel sheet with recognized students."""
    attendance_file = os.path.join(class_dir, "attendance.xlsx")
    wb = load_workbook(attendance_file)
    ws = wb.active
    today_date = datetime.today().strftime('%Y-%m-%d')

    # Check if today's date column exists, if not, add it
    if today_date not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=ws.max_column+1, value=today_date)

    # Mark attendance for recognized students
    for student in recognized_students:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == student:
                today_column_index = None
                for i, cell in enumerate(ws[1]):
                    if cell.value == today_date:
                        today_column_index = i + 1
                        break
                if today_column_index:
                    ws.cell(row=row[0].row, column=today_column_index, value="Present")
                    break

    wb.save(attendance_file)
    return attendance_file

@app.post("/mark-attendance")
async def mark_attendance(
    dept: str = Form(...),
    program: str = Form(...),
    sem: str = Form(...),
    class_: str = Form(...),
    batch_number: str = Form(...),
    file: UploadFile = File(...)
):
    try:
        # Paths for storing images and embeddings
        class_dir = f"{BASE_DIR}/batches/{batch_number}/{program}/{sem}/{class_}"
        attendance_dir = f"{class_dir}/attendance"
        os.makedirs(attendance_dir, exist_ok=True)

        # Save the uploaded file
        file_path = f"{class_dir}/{file.filename}"
        with open(file_path, "wb") as f:
            f.write(await file.read())

        # Load class embeddings
        batch_embeddings_path = f"{BASE_DIR}/batches/{batch_number}/{program}/batch_embeddings.pkl"
        if not os.path.exists(batch_embeddings_path):
            return {"message": "No embeddings found for this class"}

        with open(batch_embeddings_path, "rb") as f:
            class_embeddings = pickle.load(f)

        # Load and process the group image
        image = cv2.imread(file_path)
        faces = extract_faces(image)
        if not faces:
            return {"message": "No faces detected in the group photo"}

        recognized_students = []
        annotated_image = image.copy()

        for face_img, bbox in faces:
            # Generate face embedding
            embedding = facenet.embeddings([face_img])[0]

            # Recognize student from embeddings
            recognized_student = recognize_faces(embedding, class_embeddings)
            recognized_students.append(recognized_student)

            # Annotate the image
            x1, y1, x2, y2 = bbox
            cv2.rectangle(annotated_image, (x1, y1), (x2, y2), (0, 255, 0), 2)
            label = recognized_student if recognized_student != "Unknown" else "Unknown"
            cv2.putText(annotated_image, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (36, 255, 12), 2)

        # Save the annotated image
        annotated_image_filename = f"annotated_{file.filename}"
        annotated_image_path = os.path.join(attendance_static_dir, annotated_image_filename)
        compression_params = [cv2.IMWRITE_JPEG_QUALITY, 70]
        cv2.imwrite(annotated_image_path, annotated_image, compression_params)

        current_datetime = datetime.datetime.now()
        # Format the time in 12-hour format with AM/PM
        formatted_datetime = current_datetime.strftime("%Y-%m-%d %I:%M %p")

        attendance_meta = f"{formatted_datetime}"

        # Update the attendance Excel file
      #  attendance_file = update_attendance_excel(class_dir, recognized_students)
        print({
            "message": "Attendance marked successfully",
            "recognized_students": recognized_students,
            "annotated_image_url": f"/static/{annotated_image_filename}",
            "attendance_meta": attendance_meta
          #  "attendance_file_path": attendance_file,
        })
    
        return {
            "message": "Attendance marked successfully",
            "recognized_students": recognized_students,
            "annotated_image_url": f"/static/{annotated_image_filename}",
            "attendance_meta": attendance_meta
          #  "attendance_file_path": attendance_file,
        }
    except Exception as e:
        print("Error during attendance marking:", e)
        return {"error": str(e)}

from urllib.parse import unquote

@app.get("/attendance")
async def get_attendance(
    dept: str,
    program: str,
    sem: str,
    class_: str,
    date: str
):
    dept = unquote(dept)
    program = unquote(program)
    sem = unquote(sem)
    class_ = unquote(class_)
    date = unquote(date)

    print(f"Decoded Department: {dept}")
    print(f"Decoded Program: {program}")
    print(f"Decoded Semester: {sem}")
    print(f"Decoded Class: {class_}")
    print(f"Decoded Date: {date}")

    attendance_dir = f"{BASE_DIR}/{dept}/{program}/{sem}/{class_}/attendance"
    if not os.path.exists(attendance_dir):
        return {"message": "No attendance records found"}

    if date:
        attendance_file = f"{attendance_dir}/{date}.json"
        if os.path.exists(attendance_file):
            with open(attendance_file, "r") as f:
                return json.load(f)
        else:
            return {"message": f"No attendance record found for {date}"}
    else:
        records = []
        for filename in os.listdir(attendance_dir):
            with open(f"{attendance_dir}/{filename}", "r") as f:
                records.append(json.load(f))
        return records

